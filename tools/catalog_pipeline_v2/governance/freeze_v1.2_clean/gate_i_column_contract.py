#!/usr/bin/env python3
"""
Gate I — Column Contract Validator

Purpose:
- Validates that required columns exist in each authoritative sheet
- Prevents schema drift (columns renamed/dropped) that could break SoE consumers
- Uses logical sheet role resolution (name-based or column-based fallback)

Usage:
    python3 gate_i_column_contract.py <xlsx_file>
    
    Or run inline from series folder:
    python3 - <<'PY'
    ... (paste script content)
    PY

Rules:
- Resolves sheets by: (1) primary name match, (2) required column presence
- Each logical role has required column groups (any-one-of sets)
- Flexible naming via equivalence sets (prevents harmless variants from blocking)
"""

import openpyxl
import sys
from pathlib import Path

# ---- Helpers ----
def norm(s):
    """Normalize string for comparison: lowercase, strip, replace spaces with underscores"""
    return str(s).strip().lower().replace(" ", "_")

def sheet_headers(ws):
    """Extract normalized header set from worksheet row 1"""
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.strip():
            headers.append(norm(v))
    return set(headers)

def find_sheet_by_names(wb, names):
    """Find sheet by normalized name match"""
    want = {norm(n) for n in names}
    for s in wb.sheetnames:
        if norm(s) in want:
            return s
    return None

def find_sheet_by_required_cols(wb, required_any_sets):
    """
    Find a sheet that satisfies all required_any_sets:
    each element is a set/list of acceptable column names (any one satisfies that requirement).
    """
    for s in wb.sheetnames:
        ws = wb[s]
        hdr = sheet_headers(ws)
        ok = True
        for any_set in required_any_sets:
            any_set_norm = {norm(x) for x in any_set}
            if len(hdr.intersection(any_set_norm)) == 0:
                ok = False
                break
        if ok:
            return s
    return None

# ---- Logical roles -> primary/alt sheet names + required columns ----
ROLES = {
    "SKU_IDENTITY": {
        "sheet_names": ["NSW_L2_PRODUCTS", "NSW_SKU_MASTER_CANONICAL"],
        "required": [
            {"OEM_Catalog_No", "sku_code", "oem_catalog_no", "l2_product_code"},
            {"Make", "make"},
        ],
        "optional": [
            {"series_name", "Series", "oem_series"},
        ],
    },
    "PRICE_TRUTH": {
        "sheet_names": ["NSW_PRICE_MATRIX", "NSW_PRICE_MATRIX_CANONICAL"],
        "required": [
            {"OEM_Catalog_No", "sku_code", "oem_catalog_no", "l2_product_code"},
            {"Rate", "price", "unit_price", "rate_inr"},
            {"Currency", "currency"},
            {"EffectiveFrom", "effective_from", "wef"},
        ],
        "optional": [],
    },
    "L1_CONFIG": {
        "sheet_names": ["NSW_L1_CONFIG_LINES", "NSW_L1_LINES"],
        "required": [
            {"L1_Id", "l1_id"},
            {"L1_Group_Id", "l1_group_id"},
            {"GenericDescriptor", "generic_descriptor"},
        ],
        "optional": [
            {"capability_codes", "capability_code", "capabilities"},
            {"SC_L1_Construction", "sc_l1_construction"},
            {"SC_L2_Operation", "sc_l2_operation"},
            {"SC_L3_FeatureClass", "sc_l3_featureclass"},
            {"SC_L4_AccessoryClass", "sc_l4_accessoryclass"},
        ],
    },
    "VARIANTS": {
        "sheet_names": ["NSW_PRODUCT_VARIANTS", "NSW_VARIANTS"],
        "required": [
            {"OEM_Catalog_No", "sku_code", "oem_catalog_no", "l2_product_code"},
        ],
        "optional": [
            {"variant_code", "variant", "variant_key"},
        ],
    },
    "VARIANT_MASTER": {
        "sheet_names": ["NSW_VARIANT_MASTER"],
        "required": [
            {"variant_code", "variant", "variant_key"},
        ],
        "optional": [],
    },
}

# ---- Run ----
def main():
    if len(sys.argv) < 2:
        print("Usage: python3 gate_i_column_contract.py <xlsx_file>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    p = Path(file_path)
    
    if not p.exists():
        print(f"✗ Gate I: file not found: {p}")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(p, data_only=True)
    
    errors = []
    resolved = {}
    
    for role, cfg in ROLES.items():
        # 1) name match
        s = find_sheet_by_names(wb, cfg["sheet_names"])
        # 2) fallback by required columns
        if s is None:
            s = find_sheet_by_required_cols(wb, cfg["required"])
        
        if s is None:
            errors.append(f"{role}: missing sheet (names tried: {cfg['sheet_names']})")
            continue
        
        ws = wb[s]
        hdr = sheet_headers(ws)
        
        # verify required columns (each requirement is "any-of" set)
        missing_groups = []
        for any_set in cfg["required"]:
            any_set_norm = {norm(x) for x in any_set}
            if len(hdr.intersection(any_set_norm)) == 0:
                missing_groups.append(sorted(any_set))
        if missing_groups:
            errors.append(f"{role}: sheet '{s}' missing required column groups: {missing_groups}")
            errors.append(f"  Available headers: {sorted(hdr)}")
            continue
        
        resolved[role] = (s, sorted(hdr))
    
    # Report
    print("="*72)
    print("Gate I — Column Contract")
    print(f"File: {p}")
    print("-"*72)
    
    if errors:
        print("✗ FAIL")
        for e in errors:
            print("  -", e)
        sys.exit(1)
    
    print("✓ PASS")
    print("\nResolved sheets:")
    for role, (s, headers) in resolved.items():
        print(f"  - {role}: {s}")
        print(f"    Headers: {', '.join(headers[:10])}{'...' if len(headers) > 10 else ''}")
    
    print("="*72)

if __name__ == "__main__":
    main()

