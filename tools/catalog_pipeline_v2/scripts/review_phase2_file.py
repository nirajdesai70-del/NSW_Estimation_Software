#!/usr/bin/env python3
"""
Review Phase 2 output file - Verify TERMINOLOGY_ALIASES and README sheets
"""

from openpyxl import load_workbook
from pathlib import Path
import sys

def find_header_col(ws, header_name):
    """Find header column in row 1"""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and header_name.lower() in v.lower():
            return c
    return None

def verify_terminology_aliases(ws):
    """Verify TERMINOLOGY_ALIASES sheet"""
    print("\n" + "=" * 80)
    print("TERMINOLOGY_ALIASES VERIFICATION")
    print("=" * 80)
    
    term_col = find_header_col(ws, "term_used_in_code") or 1
    canon_col = find_header_col(ws, "canonical_term") or 2
    meaning_col = find_header_col(ws, "meaning") or 3
    
    ok_sc_to_scl = False
    bad_sc_to_cap = False
    sc_row = None
    
    print(f"\nScanning rows 2-{ws.max_row}...")
    print(f"Columns: term_used_in_code={term_col}, canonical_term={canon_col}, meaning={meaning_col}")
    
    for r in range(2, min(ws.max_row + 1, 20)):  # Check first 20 rows
        term = ws.cell(r, term_col).value
        canon = ws.cell(r, canon_col).value
        meaning = ws.cell(r, meaning_col).value if meaning_col else ""
        
        term_s = str(term).strip() if term is not None else ""
        canon_s = str(canon).strip() if canon is not None else ""
        meaning_s = str(meaning).strip() if meaning is not None else ""
        
        # Check for SC_L1..SC_L8 mapping
        if "SC_L1" in term_s.upper() and "SC_L8" in term_s.upper():
            sc_row = r
            print(f"\n  Row {r}: Found SC_L1..SC_L8 mapping")
            print(f"    term_used_in_code: {term_s}")
            print(f"    canonical_term: {canon_s}")
            print(f"    meaning: {meaning_s[:100]}..." if len(meaning_s) > 100 else f"    meaning: {meaning_s}")
            
            # Good mapping check
            if "SCL" in canon_s.upper() or "STRUCTURAL" in canon_s.upper() or "STRUCTURAL" in meaning_s.upper():
                ok_sc_to_scl = True
                print(f"    ✅ CORRECT: Maps to SCL (Structural Construction Layers)")
            
            # Bad mapping check
            if "CAPABILITY_CLASS" in canon_s.upper():
                bad_sc_to_cap = True
                print(f"    ❌ ERROR: Maps to capability_class (WRONG!)")
    
    print("\n" + "-" * 80)
    print("VERIFICATION RESULTS:")
    print(f"  ✅ SC_Lx → SCL mapping found: {ok_sc_to_scl}")
    print(f"  ❌ SC_Lx → capability_class found: {bad_sc_to_cap}")
    
    if ok_sc_to_scl and not bad_sc_to_cap:
        print("\n  ✅ TERMINOLOGY_ALIASES: PASS")
        return True
    else:
        print("\n  ❌ TERMINOLOGY_ALIASES: FAIL")
        return False

def verify_readme_sheet(ws, sheet_name):
    """Verify README sheet contains required rules"""
    print("\n" + "=" * 80)
    print(f"{sheet_name} VERIFICATION")
    print("=" * 80)
    
    # Scan first 400 rows, column 1
    blob = []
    max_r = min(ws.max_row, 400)
    print(f"\nScanning rows 1-{max_r}, column 1...")
    
    for r in range(1, max_r + 1):
        v = ws.cell(r, 1).value
        if v is not None:
            blob.append(str(v))
    
    text = "\n".join(blob).lower()
    
    has_generic = "generic naming rule" in text
    has_forcefill = "universal population rule" in text or "do not force fill" in text
    has_operating = "operating reality" in text or "operating model" in text or "layer discipline" in text
    
    print("\nContent checks:")
    print(f"  Generic Naming Rule: {'✅ FOUND' if has_generic else '❌ NOT FOUND'}")
    print(f"  Universal Population / Do Not Force Fill: {'✅ FOUND' if has_forcefill else '❌ NOT FOUND'}")
    print(f"  Operating Reality / Layer Discipline: {'✅ FOUND' if has_operating else '⚠️  NOT FOUND (warn only for README_MASTER)'}")
    
    if sheet_name == "README_ITEM_GOVERNANCE":
        if has_generic and has_forcefill:
            print("\n  ✅ README_ITEM_GOVERNANCE: PASS")
            return True
        else:
            print("\n  ❌ README_ITEM_GOVERNANCE: FAIL")
            return False
    else:  # README_MASTER
        if has_operating:
            print("\n  ✅ README_MASTER: PASS")
        else:
            print("\n  ⚠️  README_MASTER: WARNING (Operating Reality not found, but may be acceptable)")
        return True  # Not fatal for README_MASTER

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 review_phase2_file.py <path_to_phase2_file.xlsx>")
        sys.exit(1)
    
    file_path = Path(sys.argv[1])
    
    if not file_path.exists():
        print(f"Error: File not found: {file_path}")
        sys.exit(1)
    
    print("=" * 80)
    print("PHASE 2 FILE REVIEW")
    print("=" * 80)
    print(f"File: {file_path.name}")
    print(f"Path: {file_path}")
    
    try:
        wb = load_workbook(file_path, read_only=True)
        print(f"\nTotal sheets: {len(wb.sheetnames)}")
        
        # Check required sheets exist
        required_sheets = ["TERMINOLOGY_ALIASES", "README_ITEM_GOVERNANCE", "README_MASTER"]
        missing_sheets = []
        
        for sheet in required_sheets:
            if sheet in wb.sheetnames:
                print(f"  ✅ {sheet}")
            else:
                print(f"  ❌ {sheet} (MISSING)")
                missing_sheets.append(sheet)
        
        if missing_sheets:
            print(f"\n❌ ERROR: Missing required sheets: {', '.join(missing_sheets)}")
            wb.close()
            sys.exit(1)
        
        # Verify each sheet
        results = []
        
        if "TERMINOLOGY_ALIASES" in wb.sheetnames:
            results.append(verify_terminology_aliases(wb["TERMINOLOGY_ALIASES"]))
        
        if "README_ITEM_GOVERNANCE" in wb.sheetnames:
            results.append(verify_readme_sheet(wb["README_ITEM_GOVERNANCE"], "README_ITEM_GOVERNANCE"))
        
        if "README_MASTER" in wb.sheetnames:
            results.append(verify_readme_sheet(wb["README_MASTER"], "README_MASTER"))
        
        wb.close()
        
        # Final summary
        print("\n" + "=" * 80)
        print("FINAL SUMMARY")
        print("=" * 80)
        
        if all(results):
            print("\n✅ ALL VERIFICATIONS PASSED")
            print("\nPhase 2 file is correctly updated and ready for use.")
        else:
            print("\n❌ SOME VERIFICATIONS FAILED")
            print("\nPlease review the issues above and correct the file.")
            sys.exit(1)
        
    except Exception as e:
        print(f"\n❌ Error reading file: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()





