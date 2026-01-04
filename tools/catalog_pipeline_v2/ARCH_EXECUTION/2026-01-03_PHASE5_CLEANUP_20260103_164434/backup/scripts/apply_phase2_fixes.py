#!/usr/bin/env python3
"""
Phase 2: Apply fixes from normalized file to v1.4 file

This script:
1. Reads TERMINOLOGY_ALIASES from normalized file
2. Reads README rules from normalized file
3. Applies them to v1.4 file
"""

import argparse
import pandas as pd
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def read_terminology_aliases(normalized_file):
    """Read TERMINOLOGY_ALIASES sheet from normalized file."""
    try:
        df = pd.read_excel(normalized_file, sheet_name='TERMINOLOGY_ALIASES')
        print(f"✓ Read TERMINOLOGY_ALIASES from normalized file ({len(df)} rows)")
        return df
    except Exception as e:
        print(f"✗ Error reading TERMINOLOGY_ALIASES: {e}")
        return None


def read_readme_sheet(normalized_file, sheet_name):
    """Read README sheet from normalized file."""
    try:
        df = pd.read_excel(normalized_file, sheet_name=sheet_name)
        print(f"✓ Read {sheet_name} from normalized file ({len(df)} rows)")
        return df
    except Exception as e:
        print(f"✗ Error reading {sheet_name}: {e}")
        return None


def apply_terminology_aliases(v14_file, terminology_df):
    """Apply TERMINOLOGY_ALIASES fix to v1.4 file."""
    print("\n" + "=" * 80)
    print("Applying TERMINOLOGY_ALIASES fix to v1.4 file")
    print("=" * 80)
    
    try:
        # Load workbook
        wb = load_workbook(v14_file)
        
        if 'TERMINOLOGY_ALIASES' not in wb.sheetnames:
            print("⚠️  TERMINOLOGY_ALIASES sheet not found in v1.4 file")
            print("   Creating new sheet...")
            ws = wb.create_sheet('TERMINOLOGY_ALIASES')
        else:
            ws = wb['TERMINOLOGY_ALIASES']
            print(f"✓ Found TERMINOLOGY_ALIASES sheet in v1.4 file")
        
        # Clear existing content
        ws.delete_rows(1, ws.max_row)
        
        # Write headers
        headers = list(terminology_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data
        for row_idx, row_data in enumerate(terminology_df.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save
        wb.save(v14_file)
        print(f"✓ TERMINOLOGY_ALIASES applied to v1.4 file")
        return True
        
    except Exception as e:
        print(f"✗ Error applying TERMINOLOGY_ALIASES: {e}")
        return False


def apply_readme_rules(v14_file, readme_df, sheet_name):
    """Apply README rules to v1.4 file."""
    print(f"\nApplying {sheet_name} rules to v1.4 file...")
    
    try:
        # Load workbook
        wb = load_workbook(v14_file)
        
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  {sheet_name} sheet not found in v1.4 file")
            print(f"   Creating new sheet...")
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
            print(f"✓ Found {sheet_name} sheet in v1.4 file")
        
        # Clear existing content
        ws.delete_rows(1, ws.max_row)
        
        # Write headers
        headers = list(readme_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data
        for row_idx, row_data in enumerate(readme_df.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                # Handle NaN values
                if pd.isna(value):
                    value = ""
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save
        wb.save(v14_file)
        print(f"✓ {sheet_name} rules applied to v1.4 file")
        return True
        
    except Exception as e:
        print(f"✗ Error applying {sheet_name}: {e}")
        return False


def verify_consistency(normalized_file, v14_file):
    """Verify that v1.4 file matches normalized file for key sheets."""
    print("\n" + "=" * 80)
    print("Verifying Consistency")
    print("=" * 80)
    
    sheets_to_check = ['TERMINOLOGY_ALIASES', 'README_ITEM_GOVERNANCE', 'README_MASTER']
    all_match = True
    
    for sheet_name in sheets_to_check:
        try:
            df_normalized = pd.read_excel(normalized_file, sheet_name=sheet_name)
            df_v14 = pd.read_excel(v14_file, sheet_name=sheet_name)
            
            # Compare structure
            if list(df_normalized.columns) != list(df_v14.columns):
                print(f"✗ {sheet_name}: Column mismatch")
                print(f"  Normalized: {list(df_normalized.columns)}")
                print(f"  v1.4: {list(df_v14.columns)}")
                all_match = False
            elif len(df_normalized) != len(df_v14):
                print(f"⚠️  {sheet_name}: Row count mismatch ({len(df_normalized)} vs {len(df_v14)})")
            else:
                # Compare content (first few rows for TERMINOLOGY_ALIASES)
                if sheet_name == 'TERMINOLOGY_ALIASES':
                    # Check first row (the critical fix)
                    if not df_normalized.iloc[0].equals(df_v14.iloc[0]):
                        print(f"✗ {sheet_name}: First row mismatch (critical fix)")
                        print(f"  Normalized: {df_normalized.iloc[0].to_dict()}")
                        print(f"  v1.4: {df_v14.iloc[0].to_dict()}")
                        all_match = False
                    else:
                        print(f"✓ {sheet_name}: First row matches (critical fix verified)")
                else:
                    print(f"✓ {sheet_name}: Structure matches")
                    
        except Exception as e:
            print(f"✗ {sheet_name}: Error during verification - {e}")
            all_match = False
    
    return all_match


def main():
    parser = argparse.ArgumentParser(
        description='Phase 2: Apply fixes from normalized file to v1.4 file'
    )
    parser.add_argument(
        '--normalized-file',
        required=True,
        help='Path to normalized file (ITEM_Master_020126_v1.4_NORMALIZED_GAPS0.xlsx)'
    )
    parser.add_argument(
        '--v14-file',
        required=True,
        help='Path to v1.4 file (NSW_ENGINEERING_BANK_MASTER_v1.4_ACTIVE_SOP.xlsx)'
    )
    parser.add_argument(
        '--verify-only',
        action='store_true',
        help='Only verify consistency, do not apply fixes'
    )
    
    args = parser.parse_args()
    
    normalized_path = Path(args.normalized_file)
    v14_path = Path(args.v14_file)
    
    # Check files exist
    if not normalized_path.exists():
        print(f"✗ Error: Normalized file not found: {normalized_path}")
        sys.exit(1)
    
    if not v14_path.exists():
        print(f"✗ Error: v1.4 file not found: {v14_path}")
        print(f"\nPlease provide the path to NSW_ENGINEERING_BANK_MASTER_v1.4_ACTIVE_SOP.xlsx")
        sys.exit(1)
    
    print("=" * 80)
    print("Phase 2: Apply Fixes from Normalized File to v1.4 File")
    print("=" * 80)
    print(f"Normalized file: {normalized_path}")
    print(f"v1.4 file: {v14_path}")
    print()
    
    if args.verify_only:
        print("Verification mode only...")
        success = verify_consistency(normalized_path, v14_path)
        sys.exit(0 if success else 1)
    
    # Read from normalized file
    print("\n" + "=" * 80)
    print("Step 1: Reading fixes from normalized file")
    print("=" * 80)
    
    terminology_df = read_terminology_aliases(normalized_path)
    if terminology_df is None:
        print("✗ Failed to read TERMINOLOGY_ALIASES")
        sys.exit(1)
    
    readme_governance_df = read_readme_sheet(normalized_path, 'README_ITEM_GOVERNANCE')
    if readme_governance_df is None:
        print("✗ Failed to read README_ITEM_GOVERNANCE")
        sys.exit(1)
    
    readme_master_df = read_readme_sheet(normalized_path, 'README_MASTER')
    if readme_master_df is None:
        print("✗ Failed to read README_MASTER")
        sys.exit(1)
    
    # Apply fixes
    print("\n" + "=" * 80)
    print("Step 2: Applying fixes to v1.4 file")
    print("=" * 80)
    
    success1 = apply_terminology_aliases(v14_path, terminology_df)
    success2 = apply_readme_rules(v14_path, readme_governance_df, 'README_ITEM_GOVERNANCE')
    success3 = apply_readme_rules(v14_path, readme_master_df, 'README_MASTER')
    
    if not (success1 and success2 and success3):
        print("\n✗ Some fixes failed to apply")
        sys.exit(1)
    
    # Verify
    print("\n" + "=" * 80)
    print("Step 3: Verifying consistency")
    print("=" * 80)
    
    all_match = verify_consistency(normalized_path, v14_path)
    
    if all_match:
        print("\n" + "=" * 80)
        print("✅ Phase 2 Complete - All fixes applied and verified")
        print("=" * 80)
    else:
        print("\n" + "=" * 80)
        print("⚠️  Phase 2 Complete - Fixes applied but verification found issues")
        print("=" * 80)
        sys.exit(1)


if __name__ == '__main__':
    main()




