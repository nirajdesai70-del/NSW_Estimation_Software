#!/usr/bin/env python3
"""
Phase 2 (openpyxl): Apply fixes from normalized file to v1.4 file.

Copies these sheets from normalized → v1.4 target:
- TERMINOLOGY_ALIASES
- README_ITEM_GOVERNANCE
- README_MASTER

Why openpyxl:
- Preserves README layout (no pandas NaN/type issues)
- Safer for "text sheets" + merged cells + widths
- Writes to a NEW output file (no in-place overwrite)

Usage:
python3 scripts/apply_phase2_fixes_openpyxl.py \
  --normalized-file "path/ITEM_Master_020126_v1.4_NORMALIZED_GAPS0.xlsx" \
  --v14-file "path/NSW_ENGINEERING_BANK_MASTER_v1.4_ACTIVE_SOP.xlsx" \
  --out "path/NSW_ENGINEERING_BANK_MASTER_v1.4_PHASE2.xlsx"

Verify only:
python3 scripts/apply_phase2_fixes_openpyxl.py \
  --normalized-file ... \
  --v14-file ... \
  --verify-only
"""

import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from copy import copy

SHEETS_TO_COPY = [
    "TERMINOLOGY_ALIASES",
    "README_ITEM_GOVERNANCE",
    "README_MASTER",
]

def copy_dimensions(src_ws, dst_ws):
    # Column widths + hidden state
    for col, dim in src_ws.column_dimensions.items():
        dst = dst_ws.column_dimensions[col]
        dst.width = dim.width
        dst.hidden = dim.hidden
        dst.outlineLevel = dim.outlineLevel

    # Row heights + hidden state
    for r, dim in src_ws.row_dimensions.items():
        dst = dst_ws.row_dimensions[r]
        dst.height = dim.height
        dst.hidden = dim.hidden
        dst.outlineLevel = dim.outlineLevel

def clear_sheet(dst_ws):
    # Clear values (keep sheet object). Delete rows to remove old content and merges.
    if dst_ws.max_row > 1:
        dst_ws.delete_rows(1, dst_ws.max_row)
    # If sheet had 1 row, ensure it's cleared
    if dst_ws.max_row == 1 and dst_ws.max_column >= 1:
        for c in range(1, dst_ws.max_column + 1):
            dst_ws.cell(1, c).value = None

def copy_cells(src_ws, dst_ws):
    # Copy cell values + styles
    for row in src_ws.iter_rows():
        for cell in row:
            col_idx = column_index_from_string(cell.column)
            new_cell = dst_ws.cell(row=cell.row, column=col_idx, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)
            new_cell.number_format = cell.number_format
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.border = copy(cell.border)
            new_cell.alignment = copy(cell.alignment)
            new_cell.protection = copy(cell.protection)
            if cell.comment:
                new_cell.comment = copy(cell.comment)

    # Copy merged cells
    for merged in list(src_ws.merged_cells.ranges):
        dst_ws.merge_cells(str(merged))

def copy_sheet(src_wb, dst_wb, sheet_name):
    if sheet_name not in src_wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found in normalized file")

    src_ws = src_wb[sheet_name]

    # Create dst sheet if missing
    if sheet_name not in dst_wb.sheetnames:
        dst_ws = dst_wb.create_sheet(sheet_name)
    else:
        dst_ws = dst_wb[sheet_name]

    # Clear old content safely
    clear_sheet(dst_ws)

    # Copy sheet settings (basic)
    dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.sheet_view = copy(src_ws.sheet_view)
    dst_ws.page_setup = copy(src_ws.page_setup)
    dst_ws.page_margins = copy(src_ws.page_margins)

    # Copy dimensions + cells
    copy_dimensions(src_ws, dst_ws)
    copy_cells(src_ws, dst_ws)

def find_header_col(ws, header_name):
    # Find header in row 1
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.strip().lower() == header_name.strip().lower():
            return c
    return None

def verify_terminology_aliases(ws):
    """
    Semantic verification:
    - Must contain mapping SC_L1..SC_L8 -> SCL (or Structural Construction Layers)
    - Must NOT contain SC_L1..SC_L4 -> capability_class_1..4 as canonical mapping
    """
    term_col = find_header_col(ws, "term_used_in_code") or 1
    canon_col = find_header_col(ws, "canonical_term") or 2
    meaning_col = find_header_col(ws, "meaning")

    ok_sc_to_scl = False
    bad_sc_to_cap = False

    for r in range(2, ws.max_row + 1):
        term = ws.cell(r, term_col).value
        canon = ws.cell(r, canon_col).value
        meaning = ws.cell(r, meaning_col).value if meaning_col else ""

        term_s = str(term).strip() if term is not None else ""
        canon_s = str(canon).strip() if canon is not None else ""
        meaning_s = str(meaning).strip() if meaning is not None else ""

        # Good mapping check
        if term_s.upper() == "SC_L1..SC_L8" and (
            "SCL" in canon_s.upper() or "STRUCTURAL" in canon_s.upper() or "STRUCTURAL" in meaning_s.upper()
        ):
            ok_sc_to_scl = True

        # Bad mapping check
        if ("SC_L1" in term_s and "SC_L4" in term_s) and ("CAPABILITY_CLASS" in canon_s.upper()):
            bad_sc_to_cap = True

    return ok_sc_to_scl, bad_sc_to_cap

def verify_readme_contains_rules(ws):
    """
    Check for presence of the two mandatory blocks:
    - Generic Naming Rule
    - Universal Population Rule
    """
    blob = []
    max_r = min(ws.max_row, 400)  # scan first 400 lines
    for r in range(1, max_r + 1):
        v = ws.cell(r, 1).value
        if v is not None:
            blob.append(str(v))
    text = "\n".join(blob).lower()
    has_generic = "generic naming rule" in text
    has_forcefill = "universal population rule" in text or "do not force fill" in text
    has_operating = "operating reality" in text or "operating model" in text
    return has_generic, has_forcefill, has_operating

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--normalized-file", required=True)
    ap.add_argument("--v14-file", required=True)
    ap.add_argument("--out", default=None, help="Output workbook path (default: <v14>_PHASE2.xlsx)")
    ap.add_argument("--verify-only", action="store_true")
    args = ap.parse_args()

    norm = Path(args.normalized_file)
    v14 = Path(args.v14_file)

    if not norm.exists():
        raise SystemExit(f"Normalized file not found: {norm}")
    if not v14.exists():
        raise SystemExit(f"v1.4 file not found: {v14}")

    # Load workbooks
    src_wb = load_workbook(norm)
    dst_wb = load_workbook(v14)

    # If verify-only, do not modify; just check current target file status
    if args.verify_only:
        print("VERIFY ONLY MODE")
        if "TERMINOLOGY_ALIASES" in dst_wb.sheetnames:
            ok, bad = verify_terminology_aliases(dst_wb["TERMINOLOGY_ALIASES"])
            print(f"TERMINOLOGY_ALIASES: ok_sc_to_scl={ok}, bad_sc_to_cap={bad}")
        else:
            print("TERMINOLOGY_ALIASES missing in v1.4 file")

        for rn in ["README_ITEM_GOVERNANCE", "README_MASTER"]:
            if rn in dst_wb.sheetnames:
                g, f, o = verify_readme_contains_rules(dst_wb[rn])
                print(f"{rn}: generic_rule={g}, forcefill_rule={f}, operating_reality={o}")
            else:
                print(f"{rn} missing in v1.4 file")
        return

    # Apply copies
    print("APPLY MODE: Copying Phase-2 sheets from normalized → v1.4")
    for s in SHEETS_TO_COPY:
        copy_sheet(src_wb, dst_wb, s)
        print(f"  ✓ Copied sheet: {s}")

    # Verification after copy
    ok, bad = verify_terminology_aliases(dst_wb["TERMINOLOGY_ALIASES"])
    if not ok or bad:
        raise SystemExit(f"Verification failed for TERMINOLOGY_ALIASES (ok_sc_to_scl={ok}, bad_sc_to_cap={bad})")

    g1, f1, o1 = verify_readme_contains_rules(dst_wb["README_ITEM_GOVERNANCE"])
    g2, f2, o2 = verify_readme_contains_rules(dst_wb["README_MASTER"])
    if not (g1 and f1):
        raise SystemExit("README_ITEM_GOVERNANCE missing mandatory rules (Generic Naming / Do Not Force Fill).")
    # README_MASTER should contain operating reality; if not, warn (not fatal)
    if not o2:
        print("⚠️ WARNING: README_MASTER may be missing Operating Reality text. Please review.")

    # Save output
    out_path = Path(args.out) if args.out else v14.with_name(v14.stem + "_PHASE2.xlsx")
    dst_wb.save(out_path)
    print(f"\n✅ Phase 2 applied and saved: {out_path}")

if __name__ == "__main__":
    main()

